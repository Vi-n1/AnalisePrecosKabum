# -*- coding: utf-8 -*-
import pandas as pd
from pathlib import Path
from criar_dados.raspagem_kabum import Kabum
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class AnalisePrecosKabum(Kabum):
    def __init__(self):
        super().__init__()
        self.DIR_CSV = str(Path('dados/csv/Dados Kabum.csv'))
        self.DIR_XLSX = str(Path('dados/excel/Super Planilha Kabum.xlsx'))
        self.PLACAS = self.get_placas()
        self.LINKS_PLACAS = self.get_links_placas()
        self.PRECOS = self.get_precos()
        self.LINKS_IMAGENS = self.get_links_imagens()

    def cria_csv(self) -> None:
        quantidade_linhas = len(self.get_placas())
        with open(
            self.DIR_CSV,
            'w',
            encoding='utf-8',
            newline='\n',
        ) as csv_placas_precos:
            for index in range(0, quantidade_linhas):
                csv_placas_precos.write(
                    f'{self.PLACAS[index]};{self.PRECOS[index]};{self.LINKS_PLACAS[index]};{self.LINKS_IMAGENS[index]}\n'
                )
            csv_placas_precos.close()

    def aplicar_estilo_excel(self) -> None:
        planilha = load_workbook(self.DIR_XLSX)
        aba_ativa = planilha.active
        aba_ativa['A1'].fill = PatternFill(fill_type='solid', fgColor='7d7099')
        aba_ativa['B1'].fill = PatternFill(fill_type='solid', fgColor='7daba1')
        aba_ativa['C1'].fill = PatternFill(fill_type='solid', fgColor='7dab90')
        aba_ativa['D1'].fill = PatternFill(fill_type='solid', fgColor='b0c482')
        aba_ativa.column_dimensions['A'].width = 60
        aba_ativa.column_dimensions['B'].width = 11
        aba_ativa.column_dimensions['C'].width = 21
        aba_ativa.column_dimensions['D'].width = 150
        planilha.save(self.DIR_XLSX)

    def cria_excel(self) -> None:
        df = pd.read_csv(filepath_or_buffer=self.DIR_CSV, sep=';')
        df.to_excel(
            excel_writer=self.DIR_XLSX,
            engine='openpyxl',
            index=False,
        )


if __name__ == '__main__':
    app = AnalisePrecosKabum()
    app.cria_csv()
    app.cria_excel()
    app.aplicar_estilo_excel()
